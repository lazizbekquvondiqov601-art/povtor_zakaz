"""Shared Excel export helpers for all panel views."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from django.http import HttpResponse

HDR_FILL = PatternFill("solid", fgColor="1E293B")
HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
TOT_FILL = PatternFill("solid", fgColor="374151")
TOT_FONT = Font(bold=True, color="F1F5F9", size=10)
ALT_FILL = PatternFill("solid", fgColor="F8FAFC")


def to_num(s):
    """'1 234 567' → 1234567; '—' → 0."""
    try:
        return int(str(s).replace(' ', '').replace(' ', '').replace('—', '0').replace(',', ''))
    except Exception:
        return 0


def style_header(ws, row=1, fill=None, font=None):
    for cell in ws[row]:
        cell.font = font or HDR_FONT
        cell.fill = fill or HDR_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


def style_total(ws, row):
    for cell in ws[row]:
        cell.font = TOT_FONT
        cell.fill = TOT_FILL
        cell.alignment = Alignment(horizontal='right', vertical='center')
        if isinstance(cell.value, str) and cell.column == 1:
            cell.alignment = Alignment(horizontal='left', vertical='center')


def set_col_widths(ws, widths: dict):
    """widths = {col_letter: width}"""
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def make_response(wb, filename):
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    safe_name = filename.encode('ascii', errors='replace').decode()
    response['Content-Disposition'] = f'attachment; filename="{safe_name}"'
    wb.save(response)
    return response
