import sys, json
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent))

from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from django.utils import timezone
from sqlalchemy import text as sa_text
from sqlalchemy.exc import OperationalError
import pandas as pd
import openpyxl
import src.database.db_manager as db_manager
from core.dashboard_queries import (
    get_kpi_summary, get_urgent_orders, get_sales_sparkline,
    get_supplier_problems, get_category_donut,
    get_mtd_marja, get_dead_stock, get_supplier_rating,
)
from core.excel_export import style_header, style_total, set_col_widths, make_response


def _fmt(n):
    try:
        return f"{int(n):,}".replace(",", " ")
    except Exception:
        return "0"


def _fmt_mln(n):
    try:
        v = float(n) / 1_000_000
        return f"{v:.1f} mln"
    except Exception:
        return "0"


@login_required
def dashboard(request):
    today = timezone.now().strftime('%Y-%m-%d')
    engine = db_manager.engine

    # --- Asosiy ma'lumotlar ---
    spark_end = request.GET.get('spark_end', None)
    kpi              = get_kpi_summary(engine)
    urgent_orders    = get_urgent_orders(engine, limit=10)
    sales_trend      = get_sales_sparkline(engine, end_date=spark_end)
    supplier_problems = get_supplier_problems(engine, limit=5)
    category_donut   = get_category_donut(engine)

    # --- Yangi widgetlar ---
    marja_mtd       = get_mtd_marja(engine)
    dead_stock      = get_dead_stock(engine, limit=15)
    supplier_rating = get_supplier_rating(engine, limit=12)

    # --- Chart.js JSON ---
    sparkline_json = json.dumps({
        'labels': sales_trend.get('labels', []),
        'values': sales_trend.get('values', []),
    })
    donut_json = json.dumps({
        'labels': [c['kategoriya'] for c in category_donut],
        'values': [c['summa']      for c in category_donut],
    })

    # --- KPI formatlash ---
    kpi['today_sum_fmt']     = _fmt(kpi.get('today_sum', 0))
    kpi['yesterday_sum_fmt'] = _fmt(kpi.get('yesterday_sum', 0))

    # --- Marja formatlash ---
    marja_mtd['gross_profit_fmt'] = _fmt(marja_mtd['gross_profit'])
    marja_mtd['sotuv_summa_fmt']  = _fmt(marja_mtd['sotuv_summa'])
    marja_mtd['gross_mln']        = _fmt_mln(marja_mtd['gross_profit'])
    marja_mtd['sotuv_mln']        = _fmt_mln(marja_mtd['sotuv_summa'])

    # Marja rang klassi
    mp = marja_mtd['marja_pct']
    marja_mtd['color_class'] = (
        'marja-green' if mp >= 25 else
        'marja-amber' if mp >= 15 else
        'marja-red'
    )
    marja_mtd['bg_class'] = (
        'marja-bg-green' if mp >= 25 else
        'marja-bg-amber' if mp >= 15 else
        'marja-bg-red'
    )

    # --- O'lik tovarlar formatlash ---
    dead_total_summa = sum(item['jami_summa'] for item in dead_stock)
    for item in dead_stock:
        item['jami_mln'] = _fmt_mln(item['jami_summa'])

    return render(request, 'core/dashboard.html', {
        'today':            today,
        'kpi':              kpi,
        'urgent_orders':    urgent_orders,
        'sales_trend':      sales_trend,
        'supplier_problems': supplier_problems,
        'category_donut':   category_donut,
        'sparkline_json':   sparkline_json,
        'donut_json':       donut_json,
        # yangi
        'marja_mtd':        marja_mtd,
        'dead_stock':       dead_stock,
        'dead_total_mln':   _fmt_mln(dead_total_summa),
        'supplier_rating':  supplier_rating,
    })


def webapp_entry(request):
    """Telegram WebApp kirish nuqtasi. Browser orqali kirsa xato ko'rsatadi."""
    return render(request, 'core/webapp.html')


def _fmt_frozen(v):
    """Muzlagan kapital formati: >=1mln bo'lsa 'X.Y mln', aks holda raqam."""
    try:
        v = float(v)
    except Exception:
        return "0"
    if v >= 1_000_000:
        return f"{v/1_000_000:.1f} mln"
    return f"{int(v):,}".replace(",", " ")


_AGE_RANGES = [
    ('30–60 kun',   30,  60),
    ('61–90 kun',   61,  90),
    ('91–120 kun',  91, 120),
    ('121–180 kun', 121, 180),
    ('180+ kun',    181, 9999),
]

def _tovar_yoshi(imp_str):
    """DAX: IF(ISBLANK(Import_Sana), BLANK(), DATEDIFF(Import_Sana, TODAY(), DAY))"""
    try:
        from datetime import date as _d, datetime as _dt
        s = str(imp_str or '').strip()
        if not s or s in ('None', 'nan'):
            return None
        return (_d.today() - _dt.strptime(s, '%d.%m.%Y').date()).days
    except Exception:
        return None


@login_required
def olik_tovarlar(request):
    engine = db_manager.engine

    # promo_price ustuni sync da replace bilan yo'qolishi mumkin — avval qo'shamiz
    try:
        with engine.begin() as conn:
            conn.execute(sa_text('ALTER TABLE d_mahsulotlar ADD COLUMN promo_price REAL DEFAULT 0'))
    except Exception:
        pass

    # --- 1) Asosiy SQL — o'lik tovarlar ---
    # GROUP BY: Подкатегория + Артикул + Цвет(rang) → har doim 1 ta qator.
    # Bir artikul+rang uchun d_mahsulotlar da bir nechta product_id bo'lishi
    # mumkin (turli import sanasida kelgan) — agregatsiya ularni birlashtiradi:
    #   tan_narxi   = MAX(Цена поставки)
    #   sotuv_narxi = MIN(Цена продажи)
    #   qoldiq      = SUM(Кол-во)
    #   jami_summa  = SUM(Кол-во) * MAX(Цена поставки)  (muzlagan kapital)
    # f_qoldiqlar da Цена поставки/Цена продажи bo'lmasa, d_mahsulotlar dan olinadi.
    sql_main = sa_text("""
        SELECT
            COALESCE(NULLIF(TRIM(d."Подкатегория"), ''), 'Boshqa')        AS podkat,
            q."Артикул"                                                     AS artikul,
            MIN(d."Наименование")                                           AS naimenovanie,
            MAX(d."Фото")                                                    AS foto,
            MAX(d."Поставщик")                                              AS postavchik,
            MAX(d."Пол")                                                    AS pol,
            COALESCE(NULLIF(TRIM(d."Цвет"), ''), '—')                     AS rang,
            MAX(d."import_date")                                             AS import_date,
            MAX(d."promo_price")                                            AS aksiya,
            SUM(q."Кол-во")                                                AS qoldiq,
            MAX(COALESCE(q."Цена поставки", d."supply_price", 0))          AS tan_narxi,
            MIN(COALESCE(q."Цена продажи", d."Цена продажи", 0))           AS sotuv_narxi,
            SUM(q."Кол-во") * MAX(COALESCE(q."Цена поставки", d."supply_price", 0)) AS jami_summa
        FROM f_qoldiqlar q
        JOIN d_mahsulotlar d ON q.product_id = d.product_id
        WHERE date(q."Дата") = (SELECT MAX(date("Дата")) FROM f_qoldiqlar)
          AND q."Кол-во" > 0
          AND q."Артикул" NOT LIKE '010%'
          AND q."Артикул" NOT LIKE '011%'
          AND q."Наименование" NOT LIKE 'Пакет%'
        GROUP BY podkat, q."Артикул", rang
        ORDER BY jami_summa DESC
    """)
    try:
        df = pd.read_sql(sql_main, engine)
    except (OperationalError, Exception) as e:
        # Jadvallar yo'q (Railway da bot hali ma'lumot yuklamagan) — bo'sh jadval
        print(f"[olik_tovarlar] sql_main xatolik: {e}")
        df = pd.DataFrame(columns=[
            'podkat', 'artikul', 'naimenovanie', 'foto', 'postavchik', 'pol',
            'rang', 'import_date', 'aksiya', 'qoldiq', 'tan_narxi',
            'sotuv_narxi', 'jami_summa',
        ])

    # --- 2) sotuv_90 ---
    sql_s90 = sa_text("""
        SELECT s."Артикул"                                          AS artikul,
               COALESCE(NULLIF(TRIM(d."Цвет"), ''), '—')            AS rang,
               SUM(s."Продано за вычетом возвратов")                AS sotuv_90,
               SUM(s."Продажи со скидкой с учетом возвратов")       AS sotuv_summa
        FROM f_sotuvlar s
        JOIN d_mahsulotlar d ON s.product_id = d.product_id
        WHERE date(s."Дата") >= date('now', '-90 days')
        GROUP BY s."Артикул", rang
    """)
    try:
        df_s90 = pd.read_sql(sql_s90, engine)
    except (OperationalError, Exception) as e:
        print(f"[olik_tovarlar] sql_s90 xatolik: {e}")
        df_s90 = pd.DataFrame(columns=['artikul', 'rang', 'sotuv_90', 'sotuv_summa'])

    # --- Merge sotuv_90 (artikul + rang bo'yicha) ---
    if not df.empty:
        df = df.merge(df_s90, on=['artikul', 'rang'], how='left')
        df['sotuv_90']    = df['sotuv_90'].fillna(0)
        df['sotuv_summa'] = df['sotuv_summa'].fillna(0)
    else:
        df['sotuv_90']    = []
        df['sotuv_summa'] = []

    # --- OBR% hisoblash: sotuv_90 / qoldiq * 100 (turnover ratio) ---
    def _obr(row):
        try:
            q = float(row['qoldiq'])
            if q <= 0:
                return 0.0
            return round(float(row['sotuv_90']) / q * 100, 1)
        except Exception:
            return 0.0

    # --- Sotilish% (sell-through): sotuv_90 / (sotuv_90 + qoldiq) * 100 ---
    def _sotilish_foiz(row):
        try:
            s = float(row.get('sotuv_90', 0) or 0)
            q = float(row.get('qoldiq', 0) or 0)
            total = s + q
            if total <= 0:
                return 0.0
            return round(s / total * 100, 1)
        except Exception:
            return 0.0

    if not df.empty:
        df['obr'] = df.apply(_obr, axis=1)
        df['sotilish_foiz'] = df.apply(_sotilish_foiz, axis=1)
    else:
        df['obr'] = []
        df['sotilish_foiz'] = []

    # --- 3) tovar_yoshi hisoblash (DAX DATEDIFF ekvivalenti) ---
    if not df.empty:
        df['tovar_yoshi'] = df['import_date'].apply(_tovar_yoshi)
    else:
        df['tovar_yoshi'] = []

    # --- 4) Filtering OLDIN: unique qiymatlar ---
    all_podkats = sorted(df['podkat'].dropna().unique().tolist()) if not df.empty else []
    all_pols    = sorted([p for p in df['pol'].dropna().unique().tolist() if p]) if not df.empty else []

    # Qidiruv uchun unique naimenovanie ro'yxati (filtrlardan OLDIN — barcha nomlar)
    import json as _json
    all_names = sorted(df['naimenovanie'].dropna().unique().tolist()) if not df.empty else []
    names_json = _json.dumps(all_names, ensure_ascii=False)

    # --- 5) GET parametrlar ---
    selected_podkats  = request.GET.getlist('podkat')
    selected_ages     = request.GET.getlist('age')
    selected_pol      = request.GET.get('pol', '').strip()
    search_q          = request.GET.get('q', '').strip()
    only_aksiya       = request.GET.get('aksiya', '') == '1'
    with_barcode      = request.GET.get('with_barcode', '') == '1'
    sort  = request.GET.get('sort', 'jami_summa')
    order = request.GET.get('order', 'desc')
    page  = max(1, int(request.GET.get('page', 1) or 1))
    per_page = 50

    if sort not in ('qoldiq', 'jami_summa', 'sotuv_90', 'obr', 'tovar_yoshi',
                    'sotilish_foiz', 'tan_narxi', 'sotuv_narxi', 'podkat', 'artikul'):
        sort = 'jami_summa'
    if order not in ('asc', 'desc'):
        order = 'desc'

    # --- Filtering Python da ---
    if not df.empty:
        if selected_podkats:
            df = df[df['podkat'].isin(selected_podkats)]
        if selected_ages:
            mask = pd.Series(False, index=df.index)
            for age_key in selected_ages:
                for label, lo, hi in _AGE_RANGES:
                    if age_key == label:
                        yosh = df['tovar_yoshi'].fillna(0)
                        mask |= (yosh >= lo) & (yosh <= hi)
            df = df[mask]
        if only_aksiya:
            df = df[df['aksiya'].fillna(0) > 0]
        if selected_pol and 'pol' in df.columns:
            df = df[df['pol'].astype(str) == selected_pol]

    # --- Smart search (filtrlardan KEYIN, sortdan OLDIN) ---
    if search_q and not df.empty:
        q_lower = search_q.lower()
        mask = (
            df['naimenovanie'].astype(str).str.lower().str.contains(q_lower, na=False) |
            df['artikul'].astype(str).str.lower().str.contains(q_lower, na=False) |
            df['podkat'].astype(str).str.lower().str.contains(q_lower, na=False) |
            df['postavchik'].astype(str).str.lower().str.contains(q_lower, na=False)
        )
        df = df[mask]

    # --- Sort ---
    if not df.empty:
        if sort in ('podkat', 'artikul', 'naimenovanie'):
            df = df.sort_values(by=sort, ascending=(order == 'asc'),
                                key=lambda x: x.astype(str).str.lower())
        else:
            df = df.sort_values(by=sort, ascending=(order == 'asc'))

    # --- 5) Excel export ---
    if request.GET.get('export') == '1':
        if with_barcode:
            # ===== BARCODE REJIMI =====
            # podkategoriya + artikul + rang + barcode bo'yicha guruhlash:
            # har bir barcode alohida qator. sotuv_90 esa f_sotuvlar da
            # barcode bo'lmagani uchun artikul+rang bo'yicha merge qilinadi
            # (bir artikul+rang ning sotuvi barcode lar orasida takrorlanadi).
            sql_barcode = sa_text("""
                SELECT
                    COALESCE(NULLIF(TRIM(d."Подкатегория"), ''), 'Boshqa')        AS podkat,
                    q."Артикул"                                                     AS artikul,
                    MIN(d."Наименование")                                           AS naimenovanie,
                    MAX(d."Фото")                                                    AS foto,
                    COALESCE(NULLIF(TRIM(d."Цвет"), ''), '—')                     AS rang,
                    COALESCE(d."Баркод", '')                                        AS barcode,
                    MAX(d."import_date")                                             AS import_date,
                    MAX(d."promo_price")                                            AS aksiya,
                    SUM(q."Кол-во")                                                AS qoldiq,
                    MAX(COALESCE(q."Цена поставки", d."supply_price", 0))          AS tan_narxi,
                    MIN(COALESCE(q."Цена продажи", d."Цена продажи", 0))           AS sotuv_narxi,
                    SUM(q."Кол-во") * MAX(COALESCE(q."Цена поставки", d."supply_price", 0)) AS jami_summa
                FROM f_qoldiqlar q
                JOIN d_mahsulotlar d ON q.product_id = d.product_id
                WHERE date(q."Дата") = (SELECT MAX(date("Дата")) FROM f_qoldiqlar)
                  AND q."Кол-во" > 0
                  AND q."Артикул" NOT LIKE '010%'
                  AND q."Артикул" NOT LIKE '011%'
                  AND q."Наименование" NOT LIKE 'Пакет%'
                GROUP BY podkat, q."Артикул", rang, d."Баркод"
                ORDER BY jami_summa DESC
            """)
            try:
                df_bc = pd.read_sql(sql_barcode, engine)
            except (OperationalError, Exception) as e:
                print(f"[olik_tovarlar] sql_barcode xatolik: {e}")
                df_bc = pd.DataFrame(columns=[
                    'podkat', 'artikul', 'naimenovanie', 'foto', 'rang', 'barcode',
                    'import_date', 'aksiya', 'qoldiq', 'tan_narxi',
                    'sotuv_narxi', 'jami_summa',
                ])

            # --- sotuv_90 merge (artikul + rang bo'yicha) ---
            if not df_bc.empty:
                df_bc = df_bc.merge(df_s90, on=['artikul', 'rang'], how='left')
                df_bc['sotuv_90']    = df_bc['sotuv_90'].fillna(0)
                df_bc['sotuv_summa'] = df_bc['sotuv_summa'].fillna(0)
                df_bc['obr'] = df_bc.apply(_obr, axis=1)
                df_bc['sotilish_foiz'] = df_bc.apply(_sotilish_foiz, axis=1)
                df_bc['tovar_yoshi'] = df_bc['import_date'].apply(_tovar_yoshi)
            else:
                df_bc['sotuv_90']    = []
                df_bc['sotuv_summa'] = []
                df_bc['obr'] = []
                df_bc['sotilish_foiz'] = []
                df_bc['tovar_yoshi'] = []

            # --- Filtrlar (df ga emas, df_bc ga) ---
            if not df_bc.empty:
                if selected_podkats:
                    df_bc = df_bc[df_bc['podkat'].isin(selected_podkats)]
                if selected_ages:
                    mask = pd.Series(False, index=df_bc.index)
                    for age_key in selected_ages:
                        for label, lo, hi in _AGE_RANGES:
                            if age_key == label:
                                yosh = df_bc['tovar_yoshi'].fillna(0)
                                mask |= (yosh >= lo) & (yosh <= hi)
                    df_bc = df_bc[mask]
                if only_aksiya:
                    df_bc = df_bc[df_bc['aksiya'].fillna(0) > 0]

            # --- Smart search ---
            if search_q and not df_bc.empty:
                q_lower = search_q.lower()
                mask = (
                    df_bc['naimenovanie'].astype(str).str.lower().str.contains(q_lower, na=False) |
                    df_bc['artikul'].astype(str).str.lower().str.contains(q_lower, na=False) |
                    df_bc['podkat'].astype(str).str.lower().str.contains(q_lower, na=False) |
                    df_bc['postavchik'].astype(str).str.lower().str.contains(q_lower, na=False)
                )
                df_bc = df_bc[mask]

            # --- Sort ---
            if not df_bc.empty:
                if sort in ('podkat', 'artikul', 'naimenovanie'):
                    df_bc = df_bc.sort_values(by=sort, ascending=(order == 'asc'),
                                              key=lambda x: x.astype(str).str.lower())
                else:
                    df_bc = df_bc.sort_values(by=sort, ascending=(order == 'asc'))

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "O'lik tovarlar"
            ws.append([
                '#', 'Nomi', 'Podkategoriya', 'Artikul', 'Postavchik', 'Jins', 'Rang', 'Barcode', 'Aksiya',
                'Import sana', 'Tan narxi', 'Sotuv narxi', 'Sotilish%',
                'Qoldiq', 'Sotuv (dona)', 'Sotuv (summa)', 'OBR%', 'Muzlagan kapital', 'Foto URL'
            ])
            style_header(ws, row=1)

            total_qoldiq = 0
            total_s90 = 0
            total_sotuv_summa = 0
            total_jami = 0
            for i, row in enumerate(df_bc.itertuples(index=False), start=1):
                qoldiq_v      = float(getattr(row, 'qoldiq') or 0)
                s90_v         = float(getattr(row, 'sotuv_90') or 0)
                sotuv_summa_v = float(getattr(row, 'sotuv_summa') or 0)
                jami_v        = float(getattr(row, 'jami_summa') or 0)
                obr_v         = float(getattr(row, 'obr') or 0)
                tan_v         = float(getattr(row, 'tan_narxi') or 0)
                sotuv_v       = float(getattr(row, 'sotuv_narxi') or 0)
                sotilish_v    = float(getattr(row, 'sotilish_foiz') or 0)
                aksiya_v      = getattr(row, 'aksiya', None)
                aksiya_out    = '' if (aksiya_v is None or aksiya_v == 0 or
                                    (isinstance(aksiya_v, float) and aksiya_v != aksiya_v)) else int(aksiya_v)
                total_qoldiq      += qoldiq_v
                total_s90         += s90_v
                total_sotuv_summa += sotuv_summa_v
                total_jami        += jami_v
                ws.append([
                    i,
                    getattr(row, 'naimenovanie') or '',
                    getattr(row, 'podkat'),
                    getattr(row, 'artikul'),
                    getattr(row, 'postavchik', '') or '',
                    getattr(row, 'pol', '') or '',
                    getattr(row, 'rang'),
                    getattr(row, 'barcode') or '',
                    aksiya_out,
                    getattr(row, 'import_date') or '',
                    int(tan_v),
                    int(sotuv_v),
                    sotilish_v,
                    int(qoldiq_v),
                    int(s90_v),
                    int(sotuv_summa_v),
                    obr_v,
                    int(jami_v),
                    getattr(row, 'foto', '') or '',
                ])

            total_row = ws.max_row + 1
            ws.append(['', 'JAMI', '', '', '', '', '', '', '', '', '',
                       int(total_qoldiq), int(total_s90), int(total_sotuv_summa), '', int(total_jami), ''])
            style_total(ws, total_row)

            set_col_widths(ws, {
                'A': 5,  'B': 24, 'C': 18, 'D': 10, 'E': 14, 'F': 18,
                'G': 12, 'H': 14, 'I': 14, 'J': 14, 'K': 10,
                'L': 10, 'M': 12, 'N': 10, 'O': 18, 'P': 50,
            })

            return make_response(wb, "olik_barcode.xlsx")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "O'lik tovarlar"
        ws.append([
            '#', 'Nomi', 'Podkategoriya', 'Artikul', 'Postavchik', 'Jins', 'Rang', 'Aksiya',
            'Import sana', 'Tan narxi', 'Sotuv narxi', 'Sotilish%',
            'Qoldiq', 'Sotuv (dona)', 'Sotuv (summa)', 'OBR%', 'Muzlagan kapital', 'Foto URL'
        ])
        style_header(ws, row=1)

        total_qoldiq = 0
        total_s90 = 0
        total_sotuv_summa = 0
        total_jami = 0
        for i, row in enumerate(df.itertuples(index=False), start=1):
            qoldiq_v      = float(getattr(row, 'qoldiq') or 0)
            s90_v         = float(getattr(row, 'sotuv_90') or 0)
            sotuv_summa_v = float(getattr(row, 'sotuv_summa') or 0)
            jami_v        = float(getattr(row, 'jami_summa') or 0)
            obr_v         = float(getattr(row, 'obr') or 0)
            tan_v         = float(getattr(row, 'tan_narxi') or 0)
            sotuv_v       = float(getattr(row, 'sotuv_narxi') or 0)
            sotilish_v    = float(getattr(row, 'sotilish_foiz') or 0)
            aksiya_v      = getattr(row, 'aksiya', None)
            aksiya_out    = '' if (aksiya_v is None or aksiya_v == 0 or
                                (isinstance(aksiya_v, float) and aksiya_v != aksiya_v)) else int(aksiya_v)
            total_qoldiq      += qoldiq_v
            total_s90         += s90_v
            total_sotuv_summa += sotuv_summa_v
            total_jami        += jami_v
            ws.append([
                i,
                getattr(row, 'naimenovanie') or '',
                getattr(row, 'podkat'),
                getattr(row, 'artikul'),
                getattr(row, 'postavchik', '') or '',
                getattr(row, 'pol', '') or '',
                getattr(row, 'rang'),
                aksiya_out,
                getattr(row, 'import_date') or '',
                int(tan_v),
                int(sotuv_v),
                sotilish_v,
                int(qoldiq_v),
                int(s90_v),
                int(sotuv_summa_v),
                obr_v,
                int(jami_v),
                getattr(row, 'foto', '') or '',
            ])

        total_row = ws.max_row + 1
        ws.append(['', 'JAMI', '', '', '', '', '', '', '', '',
                   int(total_qoldiq), int(total_s90), int(total_sotuv_summa), '', int(total_jami), ''])
        style_total(ws, total_row)

        set_col_widths(ws, {
            'A': 5,  'B': 24, 'C': 18, 'D': 10, 'E': 14,
            'F': 12, 'G': 14, 'H': 14, 'I': 14, 'J': 10,
            'K': 10, 'L': 12, 'M': 10, 'N': 18, 'O': 50,
        })

        return make_response(wb, "olik_tovarlar.xlsx")

    # --- 6) Context uchun items tayyorlash ---
    items = []
    total_jami_raw = 0.0
    total_qoldiq_raw = 0.0
    total_s90_raw = 0.0
    total_sotuv_summa_raw = 0.0

    for r in df.itertuples(index=False):
        qoldiq_v      = float(getattr(r, 'qoldiq') or 0)
        s90_v         = float(getattr(r, 'sotuv_90') or 0)
        sotuv_summa_v = float(getattr(r, 'sotuv_summa') or 0)
        jami_v        = float(getattr(r, 'jami_summa') or 0)
        obr_v         = float(getattr(r, 'obr') or 0)
        tan_v         = float(getattr(r, 'tan_narxi') or 0)
        sotuv_v       = float(getattr(r, 'sotuv_narxi') or 0)
        sotilish_v    = float(getattr(r, 'sotilish_foiz') or 0)

        total_jami_raw        += jami_v
        total_qoldiq_raw      += qoldiq_v
        total_s90_raw         += s90_v
        total_sotuv_summa_raw += sotuv_summa_v

        if obr_v >= 70:
            obr_class = 'obr-green'
        elif obr_v >= 30:
            obr_class = 'obr-yellow'
        else:
            obr_class = 'obr-red'

        # Sotilish% (sell-through) badge: >=70 yashil, >=40 sariq, <40 qizil
        if sotilish_v >= 70:
            sotilish_class = 'obr-green'
        elif sotilish_v >= 40:
            sotilish_class = 'obr-yellow'
        else:
            sotilish_class = 'obr-red'

        yosh = getattr(r, 'tovar_yoshi', None)
        items.append({
            'podkat':         getattr(r, 'podkat'),
            'artikul':        getattr(r, 'artikul'),
            'naimenovanie':   getattr(r, 'naimenovanie') or '',
            'postavchik':     getattr(r, 'postavchik') or '',
            'pol':            getattr(r, 'pol') or '',
            'foto':           getattr(r, 'foto', None) or '',
            'rang':           getattr(r, 'rang'),
            'aksiya':         (lambda v: '' if (v is None or v == 0 or (isinstance(v, float) and v != v)) else int(v))(getattr(r, 'aksiya', None)),
            'import_date':    getattr(r, 'import_date') or '',
            'tovar_yoshi':    int(yosh) if yosh is not None else None,
            'tan_narxi':      int(tan_v),
            'sotuv_narxi':    int(sotuv_v),
            'tan_narxi_fmt':  f"{int(tan_v):,}".replace(",", " "),
            'sotuv_narxi_fmt': f"{int(sotuv_v):,}".replace(",", " "),
            'sotilish_foiz':  sotilish_v,
            'sotilish_class': sotilish_class,
            'qoldiq':         int(qoldiq_v),
            'sotuv_90':       int(s90_v),
            'sotuv_summa':    f"{int(sotuv_summa_v):,}".replace(",", " "),
            'obr':            obr_v,
            'obr_class':      obr_class,
            'jami_summa':     jami_v,
            'jami_fmt':       _fmt_frozen(jami_v),
        })

    # --- Pagination ---
    total_count = len(items)
    total_pages = max(1, (total_count + per_page - 1) // per_page)
    page = min(page, total_pages)
    items_page = items[(page - 1) * per_page : page * per_page]

    context = {
        'items':            items_page,
        'all_podkats':      all_podkats,
        'all_pols':         all_pols,
        'age_ranges':       [r[0] for r in _AGE_RANGES],
        'selected_podkats': selected_podkats,
        'selected_ages':    selected_ages,
        'selected_pol':     selected_pol,
        'search_q':         search_q,
        'names_json':       names_json,
        'only_aksiya':      only_aksiya,
        'with_barcode':     with_barcode,
        'sort':             sort,
        'order':            order,
        'total_jami':       _fmt_frozen(total_jami_raw),
        'total_count':      total_count,
        'total_qoldiq':     f"{int(total_qoldiq_raw):,}".replace(",", " "),
        'total_s90':        f"{int(total_s90_raw):,}".replace(",", " "),
        'total_sotuv_summa': f"{int(total_sotuv_summa_raw):,}".replace(",", " "),
        'has_filters':      bool(selected_podkats or selected_ages or only_aksiya or search_q or selected_pol),
        # Pagination
        'page':             page,
        'total_pages':      total_pages,
        'per_page':         per_page,
        'has_prev':         page > 1,
        'has_next':         page < total_pages,
        'page_range':       range(max(1, page - 2), min(total_pages + 1, page + 3)),
    }
    return render(request, 'core/olik.html', context)
