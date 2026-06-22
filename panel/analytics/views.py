"""
Analytics — bot ko'rinishidagi jadval:
Kategoriya | Qoldiq | Sotildi | Sotuv Summa | Foyda | OBR% | AksQoldiq | AksSot. | AksFoyda
"""
import sys, datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent))

from urllib.parse import urlencode as _urlencode

from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from django.utils import timezone
from sqlalchemy import text
from sqlalchemy.exc import OperationalError
import openpyxl
import src.database.db_manager as db_manager
from core.excel_export import to_num, style_header, style_total, set_col_widths, make_response


def _fmt(n):
    try:
        return f"{int(n):,}".replace(",", " ")
    except Exception:
        return "—"


def _get_all_kats():
    sql = text("""
        SELECT DISTINCT COALESCE(NULLIF(TRIM("Категория"), ''), 'Boshqa') AS kat
        FROM f_sotuvlar
        WHERE "Артикул" NOT LIKE '010%' AND "Артикул" NOT LIKE '011%'
          AND "Наименование" NOT LIKE 'Пакет%'
        ORDER BY kat
    """)
    session = db_manager.Session()
    try:
        return [r[0] for r in session.execute(sql).fetchall()]
    except (OperationalError, Exception) as e:
        print(f"[analytics] _get_all_kats xatolik: " + str(e).split("[SQL:")[0].rstrip())
        return []
    finally:
        session.close()


def _get_all_subkats():
    sql = text("""
        SELECT DISTINCT COALESCE(NULLIF(TRIM("Подкатегория"), ''), 'Boshqa') AS subkat
        FROM f_sotuvlar
        WHERE "Артикул" NOT LIKE '010%' AND "Артикул" NOT LIKE '011%'
          AND "Наименование" NOT LIKE 'Пакет%'
        ORDER BY subkat
    """)
    session = db_manager.Session()
    try:
        return [r[0] for r in session.execute(sql).fetchall()]
    except (OperationalError, Exception) as e:
        print(f"[analytics] _get_all_subkats xatolik: " + str(e).split("[SQL:")[0].rstrip())
        return []
    finally:
        session.close()


def _build_table(start_date, end_date, kat_filters=None, subkat_filters=None):
    params = {"s": start_date, "e": end_date}

    kat_where = subkat_where = ""
    q_kat_where = q_subkat_where = ""  # f_qoldiqlar uchun (d_mahsulotlar JOIN bilan)

    if kat_filters:
        placeholders = ', '.join(f':k{i}' for i in range(len(kat_filters)))
        kat_where   = f'AND "Категория" IN ({placeholders})'
        q_kat_where = f'AND d."Категория" IN ({placeholders})'
        for i, v in enumerate(kat_filters):
            params[f'k{i}'] = v
        grp   = '"Подкатегория"'
        q_grp = 'd."Подкатегория"'
    else:
        grp   = '"Категория"'
        q_grp = 'q."Категория"'

    if subkat_filters:
        placeholders = ', '.join(f':sk{i}' for i in range(len(subkat_filters)))
        subkat_where   = f'AND "Подкатегория" IN ({placeholders})'
        q_subkat_where = f'AND d."Подкатегория" IN ({placeholders})'
        for i, v in enumerate(subkat_filters):
            params[f'sk{i}'] = v

    label   = f'COALESCE(NULLIF(TRIM({grp}), \'\'), \'Boshqa\')'
    q_label = f'COALESCE(NULLIF(TRIM({q_grp}), \'\'), \'Boshqa\')'

    sql_asosiy = text(f"""
        SELECT
            {label} AS kat,
            SUM("Продано за вычетом возвратов")                AS sotildi,
            SUM("Продажи со скидкой с учетом возвратов")       AS summa,
            SUM("Валовая прибыль")                              AS foyda
        FROM f_sotuvlar
        WHERE date("Дата") >= :s AND date("Дата") <= :e
          AND "Артикул" NOT LIKE '010%'
          AND "Артикул" NOT LIKE '011%'
          AND "Наименование" NOT LIKE 'Пакет%'
          {kat_where} {subkat_where}
        GROUP BY kat
    """)

    sql_aksiya = text(f"""
        SELECT
            {label} AS kat,
            SUM("Продано за вычетом возвратов")                AS sotildi,
            SUM("Продажи со скидкой с учетом возвратов")       AS summa,
            SUM("Валовая прибыль")                              AS foyda
        FROM f_sotuvlar
        WHERE date("Дата") >= :s AND date("Дата") <= :e
          AND ("Артикул" LIKE '010%' OR "Артикул" LIKE '011%')
          AND "Наименование" NOT LIKE 'Пакет%'
          {kat_where} {subkat_where}
        GROUP BY kat
    """)

    # Qoldiq — filter end_date gacha bo'lgan eng oxirgi mavjud sana
    # MAX(Дата) emas, balki end_date dan katta bo'lmagan eng oxirgi sana olinadi.
    # Masalan: filter 01.06–19.06 → 19.06 qoldig'i, 20.06 emas.
    sql_q_as = text(f"""
        SELECT {q_label} AS kat, SUM(q."Кол-во") AS qty
        FROM f_qoldiqlar q
        JOIN d_mahsulotlar d ON q.product_id = d.product_id
        WHERE date(q."Дата") = (
            SELECT MAX(date("Дата")) FROM f_qoldiqlar WHERE date("Дата") <= :e
        )
          AND q."Артикул" NOT LIKE '010%' AND q."Артикул" NOT LIKE '011%'
          AND TRIM(COALESCE(q."Категория",'')) != ''
          {q_kat_where} {q_subkat_where}
        GROUP BY kat HAVING SUM(q."Кол-во") > 0
    """)
    sql_q_ak = text(f"""
        SELECT {q_label} AS kat, SUM(q."Кол-во") AS qty
        FROM f_qoldiqlar q
        JOIN d_mahsulotlar d ON q.product_id = d.product_id
        WHERE date(q."Дата") = (
            SELECT MAX(date("Дата")) FROM f_qoldiqlar WHERE date("Дата") <= :e
        )
          AND (q."Артикул" LIKE '010%' OR q."Артикул" LIKE '011%')
          AND TRIM(COALESCE(q."Категория",'')) != ''
          {q_kat_where} {q_subkat_where}
        GROUP BY kat HAVING SUM(q."Кол-во") > 0
    """)

    # O'rtacha qoldiq — DAX "Средний Остаток Python" formulasi:
    # Har bir DO'KON uchun o'rtacha kunlik qoldiq hisoblanib, keyin do'konlar yig'iladi.
    # SUMX(shops, AVG_per_shop(daily_qty)) — oddiy AVG(daily_total) dan farqli.
    sql_avg_as = text(f"""
        SELECT kat, SUM(shop_avg) AS avg_qty FROM (
            SELECT kat, q_mag, AVG(daily_qty) AS shop_avg FROM (
                SELECT {q_label}    AS kat,
                       q."Магазин"  AS q_mag,
                       date(q."Дата") AS dt,
                       SUM(q."Кол-во") AS daily_qty
                FROM f_qoldiqlar q
                JOIN d_mahsulotlar d ON q.product_id = d.product_id
                WHERE date(q."Дата") >= :s AND date(q."Дата") <= :e
                  AND q."Артикул" NOT LIKE '010%' AND q."Артикул" NOT LIKE '011%'
                  AND TRIM(COALESCE(q."Категория",'')) != ''
                  {q_kat_where} {q_subkat_where}
                GROUP BY kat, q_mag, dt
            ) t1 GROUP BY kat, q_mag
        ) t2 GROUP BY kat
    """)
    sql_avg_ak = text(f"""
        SELECT kat, SUM(shop_avg) AS avg_qty FROM (
            SELECT kat, q_mag, AVG(daily_qty) AS shop_avg FROM (
                SELECT {q_label}    AS kat,
                       q."Магазин"  AS q_mag,
                       date(q."Дата") AS dt,
                       SUM(q."Кол-во") AS daily_qty
                FROM f_qoldiqlar q
                JOIN d_mahsulotlar d ON q.product_id = d.product_id
                WHERE date(q."Дата") >= :s AND date(q."Дата") <= :e
                  AND (q."Артикул" LIKE '010%' OR q."Артикул" LIKE '011%')
                  AND TRIM(COALESCE(q."Категория",'')) != ''
                  {q_kat_where} {q_subkat_where}
                GROUP BY kat, q_mag, dt
            ) t1 GROUP BY kat, q_mag
        ) t2 GROUP BY kat
    """)

    # Marja — 010/011 ham kiritiladi, faqat sana va kat/subkat filtr ishlaydi
    sql_marja = text(f"""
        SELECT
            {label} AS kat,
            SUM("Валовая прибыль")                            AS foyda_all,
            SUM("Продажи со скидкой с учетом возвратов")      AS summa_all
        FROM f_sotuvlar
        WHERE date("Дата") >= :s AND date("Дата") <= :e
          AND "Наименование" NOT LIKE 'Пакет%'
          {kat_where} {subkat_where}
        GROUP BY kat
    """)

    session = db_manager.Session()
    try:
        as_rows   = session.execute(sql_asosiy, params).fetchall()
        ak_rows   = session.execute(sql_aksiya, params).fetchall()
        q_as      = {r[0]: float(r[1] or 0) for r in session.execute(sql_q_as,    params).fetchall()}
        q_ak      = {r[0]: float(r[1] or 0) for r in session.execute(sql_q_ak,    params).fetchall()}
        avg_as    = {r[0]: float(r[1] or 0) for r in session.execute(sql_avg_as,  params).fetchall()}
        marja_raw = {r[0]: (float(r[1] or 0), float(r[2] or 0))
                     for r in session.execute(sql_marja, params).fetchall()}
    except (OperationalError, Exception) as e:
        # Jadvallar yo'q (Railway da bot hali ma'lumot yuklamagan) — bo'sh natija
        print(f"[analytics] _build_table xatolik: " + str(e).split("[SQL:")[0].rstrip())
        as_rows, ak_rows = [], []
        q_as, q_ak, avg_as, marja_raw = {}, {}, {}, {}
    finally:
        session.close()

    asosiy = {r[0]: {"sotildi": float(r[1] or 0), "summa": float(r[2] or 0), "foyda": float(r[3] or 0)}
              for r in as_rows}
    aksiya  = {r[0]: {"sotildi": float(r[1] or 0), "summa": float(r[2] or 0), "foyda": float(r[3] or 0)}
              for r in ak_rows}

    all_kats = sorted(set(asosiy) | set(q_as))

    rows = []
    for kat in all_kats:
        a  = asosiy.get(kat, {"sotildi": 0, "summa": 0, "foyda": 0})
        ak = aksiya.get(kat,  {"sotildi": 0, "summa": 0, "foyda": 0})
        qoldiq      = q_as.get(kat, 0)
        aks_qoldiq  = q_ak.get(kat, 0)
        avg_qoldiq  = avg_as.get(kat, 0)
        obr = round(a["sotildi"] / avg_qoldiq * 100) if avg_qoldiq > 0 else 0
        m_foyda, m_summa = marja_raw.get(kat, (0, 0))
        marja = min(round(m_foyda / m_summa * 100, 1), 100.0) if m_summa > 0 else 0

        rows.append({
            "kat":        kat,
            "qoldiq":     int(qoldiq),
            "sotildi":    int(a["sotildi"]),
            "summa":      int(a["summa"]),
            "foyda":      int(a["foyda"]),
            "marja":      marja,
            "obr":        obr,
            "aks_qoldiq": int(aks_qoldiq),
            "aks_sotildi":int(ak["sotildi"]),
            "aks_foyda":  int(ak["foyda"]),
        })

    rows.sort(key=lambda x: x["foyda"], reverse=True)

    for r in rows:
        r["qoldiq"]      = _fmt(r["qoldiq"])
        r["sotildi"]     = _fmt(r["sotildi"])
        r["summa"]       = _fmt(r["summa"])
        r["foyda"]       = _fmt(r["foyda"])
        r["aks_qoldiq"]  = _fmt(r["aks_qoldiq"])
        r["aks_sotildi"] = _fmt(r["aks_sotildi"])
        r["aks_foyda"]   = _fmt(r["aks_foyda"])

    total_m_foyda = sum(v[0] for v in marja_raw.values())
    total_m_summa = sum(v[1] for v in marja_raw.values())

    totals = {
        "qoldiq":      int(sum(q_as.values())),
        "sotildi":     int(sum(a["sotildi"] for a in asosiy.values())),
        "summa":       int(sum(a["summa"]   for a in asosiy.values())),
        "foyda":       int(sum(a["foyda"]   for a in asosiy.values())),
        "marja":       min(round(total_m_foyda / total_m_summa * 100, 1), 100.0) if total_m_summa > 0 else 0,
        "aks_qoldiq":  int(sum(q_ak.values())),
        "aks_sotildi": int(sum(a["sotildi"] for a in aksiya.values())),
        "aks_foyda":   int(sum(a["foyda"]   for a in aksiya.values())),
    }
    avg_tot = sum(avg_as.values())
    totals["obr"] = round(totals["sotildi"] / avg_tot * 100) if avg_tot > 0 else 0

    totals["qoldiq"]      = _fmt(totals["qoldiq"])
    totals["sotildi"]     = _fmt(totals["sotildi"])
    totals["summa"]       = _fmt(totals["summa"])
    totals["foyda"]       = _fmt(totals["foyda"])
    totals["aks_qoldiq"]  = _fmt(totals["aks_qoldiq"])
    totals["aks_sotildi"] = _fmt(totals["aks_sotildi"])
    totals["aks_foyda"]   = _fmt(totals["aks_foyda"])

    return rows, totals


def _analytics_excel(rows, totals, start_date, end_date):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sotuv Tahlili"
    ws.freeze_panes = "B2"

    headers = ['Kategoriya', 'Qoldiq', 'Sotildi', 'Sot.Summa', 'Foyda', 'Marja%', 'OBR%',
               'AksQoldiq', 'AksSot.', 'AksFoyda']
    ws.append(headers)
    style_header(ws, 1)
    ws.row_dimensions[1].height = 28

    for r in rows:
        ws.append([
            r['kat'],
            to_num(r['qoldiq']),
            to_num(r['sotildi']),
            to_num(r['summa']),
            to_num(r['foyda']),
            r['marja'] / 100,
            r['obr'] / 100,
            to_num(r['aks_qoldiq']),
            to_num(r['aks_sotildi']),
            to_num(r['aks_foyda']),
        ])

    ws.append([
        'JAMI',
        to_num(totals['qoldiq']),
        to_num(totals['sotildi']),
        to_num(totals['summa']),
        to_num(totals['foyda']),
        totals['marja'] / 100,
        totals['obr'] / 100,
        to_num(totals['aks_qoldiq']),
        to_num(totals['aks_sotildi']),
        to_num(totals['aks_foyda']),
    ])
    style_total(ws, ws.max_row)

    # Format percent columns
    for row in ws.iter_rows(min_row=2, min_col=6, max_col=7):
        for cell in row:
            cell.number_format = '0.0%'
    # Format number columns
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=5):
        for cell in row:
            cell.number_format = '#,##0'
    for row in ws.iter_rows(min_row=2, min_col=8, max_col=10):
        for cell in row:
            cell.number_format = '#,##0'

    set_col_widths(ws, {'A': 26, 'B': 12, 'C': 12, 'D': 14, 'E': 14,
                        'F': 10, 'G': 10, 'H': 12, 'I': 12, 'J': 14})
    return make_response(wb, f"analytics_{start_date}_{end_date}.xlsx")


@login_required
def analytics_main(request):
    today      = timezone.now().date()
    start_date = request.GET.get("start", str(today.replace(day=1)))
    end_date   = request.GET.get("end",   str(today))
    kat_filters    = [k.strip() for k in request.GET.getlist("kat") if k.strip()]
    subkat_filters = [s.strip() for s in request.GET.getlist("subkat") if s.strip()]

    rows, totals = _build_table(
        start_date, end_date,
        kat_filters    or None,
        subkat_filters or None,
    )

    if request.GET.get('export') == '1':
        return _analytics_excel(rows, totals, start_date, end_date)

    qoldiq_sana = db_manager.get_max_stock_date_str()
    all_kats    = _get_all_kats()
    all_subkats = _get_all_subkats()

    first_this       = today.replace(day=1)
    last_month_end   = first_this - datetime.timedelta(days=1)
    last_month_start = last_month_end.replace(day=1)

    col_label = "Podkategoriya" if kat_filters else "Kategoriya"

    _p = [('start', start_date), ('end', end_date)]
    for k in kat_filters:    _p.append(('kat', k))
    for s in subkat_filters: _p.append(('subkat', s))
    _p.append(('export', '1'))
    export_url = '/analytics/?' + _urlencode(_p)

    return render(request, "analytics/main.html", {
        "start_date": start_date,
        "end_date":   end_date,
        "today":      str(today),
        "oy_boshi":   str(first_this),
        "otgan_oy_start": str(last_month_start),
        "otgan_oy_end":   str(last_month_end),
        "qoldiq_sana":    qoldiq_sana,
        "rows":           rows,
        "totals":         totals,
        "kat_filters":    kat_filters,
        "subkat_filters": subkat_filters,
        "all_kats":       all_kats,
        "all_subkats":    all_subkats,
        "col_label":      col_label,
        "export_url":     export_url,
    })
