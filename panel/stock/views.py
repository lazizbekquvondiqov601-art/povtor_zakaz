"""
Stock app — views.
Qoldiqlar: sana va kategoriya bo'yicha filtrlash + kunlik grafik.
"""
import sys, json
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent))

from django.contrib.auth.decorators import login_required
from django.shortcuts import render
import openpyxl
import pandas as pd
import src.database.db_manager as db_manager
from core.excel_export import style_header, style_total, set_col_widths, make_response


@login_required
def stock_main(request):
    dates = db_manager.get_last_7_stock_dates()
    selected_date     = request.GET.get('date', dates[0] if dates else '')
    selected_category = request.GET.get('category', '')

    categories = []
    if selected_date:
        categories = db_manager.get_stock_categories_on_date(selected_date)

    asosiy = []
    aksiya = []
    total_asosiy = 0
    total_aksiya = 0
    total_all    = 0
    chart_json   = 'null'

    if selected_date and selected_category:
        asosiy, aksiya = db_manager.get_stock_subcategories_summary_v2(
            selected_category, selected_date
        )
        total_asosiy = int(sum(item[2] for item in asosiy))
        total_aksiya = int(sum(item[2] for item in aksiya))
        total_all    = total_asosiy + total_aksiya

        # Kunlik sotuv (oxirgi 30 kun, tanlangan kategoriya)
        try:
            sotuv_df = pd.read_sql('''
                SELECT date("Дата") as kun,
                       SUM("Продано за вычетом возвратов") as sotildi
                FROM f_sotuvlar
                WHERE "Категория" = :cat
                GROUP BY kun ORDER BY kun
            ''', db_manager.engine, params={'cat': selected_category})

            # Asosiy qoldiq kunlik (010/011 dan tashqari)
            qoldiq_df = pd.read_sql('''
                SELECT date("Дата") as kun,
                       SUM("Кол-во") as qoldiq
                FROM f_qoldiqlar
                WHERE "Категория" = :cat
                  AND "Артикул" NOT LIKE '010%'
                  AND "Артикул" NOT LIKE '011%'
                GROUP BY kun ORDER BY kun
            ''', db_manager.engine, params={'cat': selected_category})

            # Birlashtirish
            merged = pd.merge(sotuv_df, qoldiq_df, on='kun', how='outer').sort_values('kun')
            merged['sotildi'] = pd.to_numeric(merged['sotildi'], errors='coerce').fillna(0)
            merged['qoldiq']  = pd.to_numeric(merged['qoldiq'],  errors='coerce').fillna(0)

            chart_json = json.dumps({
                'labels':  merged['kun'].tolist(),
                'sotuv':   [round(float(v), 0) for v in merged['sotildi']],
                'qoldiq':  [round(float(v), 0) for v in merged['qoldiq']],
            })
        except Exception:
            chart_json = 'null'

    if request.GET.get('export') == '1' and selected_date and selected_category:
        wb = openpyxl.Workbook()

        ws1 = wb.active
        ws1.title = "Asosiy"
        ws1.append(['Podkategoriya', 'Qoldiq (dona)'])
        style_header(ws1, 1)
        for item in asosiy:
            ws1.append([item[1], int(item[2])])
        ws1.append(['JAMI', total_asosiy])
        style_total(ws1, ws1.max_row)
        set_col_widths(ws1, {'A': 30, 'B': 16})

        ws2 = wb.create_sheet("Aksiya (010-011)")
        ws2.append(['Podkategoriya', 'Qoldiq (dona)'])
        style_header(ws2, 1)
        for item in aksiya:
            ws2.append([item[1], int(item[2])])
        ws2.append(['JAMI', total_aksiya])
        style_total(ws2, ws2.max_row)
        set_col_widths(ws2, {'A': 30, 'B': 16})

        safe_cat = selected_category.replace('/', '_')
        return make_response(wb, f"stock_{selected_date}_{safe_cat}.xlsx")

    return render(request, 'stock/main.html', {
        'dates':             dates,
        'selected_date':     selected_date,
        'categories':        categories,
        'selected_category': selected_category,
        'asosiy':            asosiy,
        'aksiya':            aksiya,
        'total_asosiy':      total_asosiy,
        'total_aksiya':      total_aksiya,
        'total_all':         total_all,
        'chart_json':        chart_json,
    })
