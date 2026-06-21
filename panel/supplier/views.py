"""
Supplier app — views.
Supplierlar ro'yxati (oy boshidan statistika) + detail (sotuv/qoldiq + grafik).
"""
import sys, json
from pathlib import Path
from datetime import date

sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent))

from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.shortcuts import render
from sqlalchemy import text as sa_text
import pandas as pd
import openpyxl
import src.database.db_manager as db_manager
from core.excel_export import style_header, style_total, set_col_widths, make_response


def _fmt(val, decimals=0):
    try:
        return f"{float(val):,.{decimals}f}".replace(",", " ")
    except Exception:
        return "0"


def _obr_class(obr):
    if obr >= 99:
        return 'obr-green'
    elif obr >= 85:
        return 'obr-yellow'
    elif obr >= 70:
        return 'obr-orange'
    else:
        return 'obr-red'


def _marja_class(marja):
    if marja >= 25:
        return 'marja-green'
    elif marja >= 15:
        return 'marja-yellow'
    else:
        return 'marja-red'


def _get_avg_qoldiq(oy_boshi):
    """Oy boshidan bugunga qadar har bir supplier uchun o'rtacha kunlik qoldiqni hisoblaydi."""
    sql = """
        SELECT supplier, AVG(daily_qty) AS avg_qoldiq
        FROM (
            SELECT
                d."Поставщик" AS supplier,
                date(q."Дата") AS dt,
                SUM(q."Кол-во") AS daily_qty
            FROM f_qoldiqlar q
            JOIN d_mahsulotlar d ON q.product_id = d.product_id
            WHERE date(q."Дата") >= :oy_boshi
              AND d."Поставщик" IS NOT NULL
              AND d."Поставщик" != ''
              AND q."Артикул" NOT LIKE '010%'
              AND q."Артикул" NOT LIKE '011%'
            GROUP BY d."Поставщик", dt
        ) t
        GROUP BY supplier
    """
    try:
        df = pd.read_sql(sql, db_manager.engine, params={'oy_boshi': oy_boshi})
        return {row['supplier']: float(row['avg_qoldiq']) for _, row in df.iterrows()}
    except Exception:
        return {}


@login_required
def supplier_list(request):
    query = request.GET.get('q', '').strip()

    # Oy boshi
    oy_boshi = str(date.today().replace(day=1))

    # O'rtacha kunlik qoldiq (OBR hisoblash uchun)
    avg_qoldiq_map = _get_avg_qoldiq(oy_boshi)

    # Statistika bilan barcha supplierlar
    df = db_manager.get_all_suppliers_stats()

    if df.empty:
        # Fallback: statistiksiz ro'yxat
        try:
            raw = pd.read_sql(
                'SELECT DISTINCT "Поставщик" FROM d_mahsulotlar '
                'WHERE "Поставщик" IS NOT NULL AND "Поставщик" != \'\' '
                'ORDER BY "Поставщик"',
                db_manager.engine
            )
            suppliers = [
                {
                    'name': s,
                    'rank': idx + 1,
                    'sotildi': 0,
                    'summa': '0',
                    'foyda': '0',
                    'qoldiq': 0,
                    'foyda_raw': 0.0,
                    'summa_raw': 0.0,
                    'obr': 0.0,
                    'obr_class': _obr_class(0),
                    'marja': 0.0,
                    'marja_class': _marja_class(0),
                }
                for idx, s in enumerate(raw['Поставщик'].dropna().tolist()) if s
            ]
        except Exception:
            suppliers = []
        has_stats = False
    else:
        has_stats = True
        suppliers = []
        for i, row in df.iterrows():
            name = str(row['supplier']).strip()
            if not name:
                continue

            sotildi_val = float(row.get('sotildi', 0))
            foyda_raw   = float(row.get('foyda', 0))
            summa_raw   = float(row.get('summa', 0))

            # OBR% hisoblash
            avg_q = avg_qoldiq_map.get(name, 0.0)
            obr_val = round(sotildi_val / avg_q * 100, 1) if avg_q > 0 else 0.0
            obr_val = min(obr_val, 9999.0)  # cheksiz raqamlarni cheklash

            # Marja% hisoblash
            marja_val = min(round(foyda_raw / summa_raw * 100, 1), 100.0) if summa_raw > 0 else 0.0

            suppliers.append({
                'name':        name,
                'rank':        i + 1,
                'sotildi':     int(sotildi_val),
                'summa':       _fmt(summa_raw),
                'foyda':       _fmt(foyda_raw),
                'qoldiq':      int(row.get('qoldiq', 0)),
                'foyda_raw':   foyda_raw,
                'summa_raw':   summa_raw,
                'obr':         obr_val,
                'obr_class':   _obr_class(obr_val),
                'marja':       marja_val,
                'marja_class': _marja_class(marja_val),
            })

    # Qidiruv
    if query:
        suppliers = [s for s in suppliers if query.lower() in s['name'].lower()]

    # Excel export
    if request.GET.get('export') == '1':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Supplierlar"
        ws.freeze_panes = "B2"
        ws.append(['#', 'Supplier', 'OBR%', 'Qoldiq', 'Sotildi', 'Sotuv summa', 'Foyda', 'Marja%'])
        style_header(ws, 1)
        for s in suppliers:
            ws.append([
                s.get('rank', ''),
                s['name'],
                s.get('obr', 0.0),
                s.get('qoldiq', 0),
                s.get('sotildi', 0),
                s.get('summa', '0'),
                s.get('foyda', '0'),
                s.get('marja', 0.0),
            ])
        set_col_widths(ws, {
            'A': 5,
            'B': 30,
            'C': 10,
            'D': 14,
            'E': 14,
            'F': 18,
            'G': 18,
            'H': 10,
        })
        return make_response(wb, "suppliers.xlsx")

    # Pagination
    per_page = 25
    paginator = Paginator(suppliers, per_page)
    page_obj = paginator.get_page(request.GET.get('page', 1))

    return render(request, 'supplier/list.html', {
        'page_obj':  page_obj,
        'query':     query,
        'total':     len(suppliers),
        'has_stats': has_stats,
    })


@login_required
def supplier_detail(request, supplier_name):
    sotuv, qoldiq, start_date, end_date = db_manager.get_supplier_sales_report(supplier_name)

    # Artikul soni
    try:
        cnt = pd.read_sql(
            'SELECT COUNT(DISTINCT "Артикул") as cnt FROM d_mahsulotlar WHERE "Поставщик" = :s',
            db_manager.engine, params={'s': supplier_name}
        )
        product_count = int(cnt['cnt'].iloc[0]) if not cnt.empty else 0
    except Exception:
        product_count = 0

    # Kategoriya bo'yicha sotuv summa (Marja%) va avg_qoldiq (OBR%)
    summa_by_kat = {}
    avg_q_by_kat = {}
    if start_date:
        like_pat = f"%{supplier_name}%"
        try:
            summa_df = pd.read_sql(sa_text('''
                SELECT
                    COALESCE(NULLIF(TRIM(s."Категория"), ''), 'Boshqa') AS kat,
                    SUM(s."Продажи со скидкой с учетом возвратов") AS summa
                FROM f_sotuvlar s
                JOIN d_mahsulotlar d ON s.product_id = d.product_id
                WHERE d."Поставщик" LIKE :sup
                  AND date(s."Дата") >= :sd
                  AND s."Артикул" NOT LIKE '010%'
                  AND s."Артикул" NOT LIKE '011%'
                  AND s."Наименование" NOT LIKE 'Пакет%'
                GROUP BY kat
            '''), db_manager.engine, params={'sup': like_pat, 'sd': start_date})
            summa_by_kat = {r['kat']: float(r['summa'] or 0) for _, r in summa_df.iterrows()}
        except Exception:
            pass

        try:
            avgq_df = pd.read_sql(sa_text('''
                SELECT kat, AVG(daily_qty) AS avg_qoldiq
                FROM (
                    SELECT
                        COALESCE(NULLIF(TRIM(q."Категория"), ''), 'Boshqa') AS kat,
                        date(q."Дата") AS dt,
                        SUM(q."Кол-во") AS daily_qty
                    FROM f_qoldiqlar q
                    JOIN d_mahsulotlar d ON q.product_id = d.product_id
                    WHERE d."Поставщик" LIKE :sup
                      AND date(q."Дата") >= :sd
                      AND q."Артикул" NOT LIKE '010%'
                      AND q."Артикул" NOT LIKE '011%'
                    GROUP BY kat, dt
                ) t
                GROUP BY kat
            '''), db_manager.engine, params={'sup': like_pat, 'sd': start_date})
            avg_q_by_kat = {r['kat']: float(r['avg_qoldiq'] or 0) for _, r in avgq_df.iterrows()}
        except Exception:
            pass

    # Barcha kategoriyalarni birlashtir
    all_kats = set(list(sotuv.keys()) + list(qoldiq.keys()))
    total_summa_raw = 0.0
    kategoriyalar = []
    for kat in all_kats:
        sotildi_val = float(sotuv.get(kat, {}).get('qty', 0))
        foyda_val   = float(sotuv.get(kat, {}).get('profit', 0))
        qoldiq_val  = float(qoldiq.get(kat, 0))
        summa_val   = summa_by_kat.get(kat, 0)
        avg_q       = avg_q_by_kat.get(kat, 0)

        obr   = min(round(sotildi_val / avg_q * 100, 1), 9999.0) if avg_q > 0 else 0.0
        marja = min(round(foyda_val / summa_val * 100, 1), 100.0) if summa_val > 0 else 0.0

        total_summa_raw += summa_val
        kategoriyalar.append({
            'kat':        kat,
            'sotildi':    int(sotildi_val),
            'qoldiq':     int(qoldiq_val),
            'summa':      _fmt(summa_val),
            'foyda':      _fmt(foyda_val),
            'foyda_raw':  foyda_val,
            'obr':        obr,
            'obr_class':  _obr_class(obr),
            'marja':      marja,
            'marja_class': _marja_class(marja),
        })
    kategoriyalar.sort(key=lambda x: x['foyda_raw'], reverse=True)

    # KPI jami
    total_qty    = sum(v.get('qty', 0)    for v in sotuv.values())
    total_profit = sum(v.get('profit', 0) for v in sotuv.values())
    total_qoldiq = sum(qoldiq.values())

    # Kunlik grafik
    chart_data = db_manager.get_supplier_daily_chart(supplier_name)
    chart_json = json.dumps(chart_data) if chart_data else 'null'

    if request.GET.get('export') == '1':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Kategoriyalar"
        ws.freeze_panes = "B2"
        ws.append(['#', 'Kategoriya', 'Sotildi', 'Qoldiq', 'Sotuv summa', 'Foyda', 'OBR%', 'Marja%'])
        style_header(ws, 1)
        for i, k in enumerate(kategoriyalar, 1):
            ws.append([i, k['kat'], k['sotildi'], k['qoldiq'], k['summa'], k['foyda'], k['obr'], k['marja']])
        ws.append(['', 'JAMI', int(total_qty), int(total_qoldiq), _fmt(total_summa_raw), _fmt(total_profit), '', ''])
        style_total(ws, ws.max_row)
        set_col_widths(ws, {'A': 5, 'B': 30, 'C': 12, 'D': 12, 'E': 18, 'F': 18, 'G': 10, 'H': 10})
        safe = supplier_name.replace('/', '_').replace('\\', '_')
        return make_response(wb, f"supplier_{safe}.xlsx")

    return render(request, 'supplier/detail.html', {
        'supplier_name':  supplier_name,
        'kategoriyalar':  kategoriyalar,
        'start_date':     start_date,
        'end_date':       end_date,
        'product_count':  product_count,
        'total_qty':      int(total_qty),
        'total_profit':   _fmt(total_profit),
        'total_qoldiq':   int(total_qoldiq),
        'total_summa':    _fmt(total_summa_raw),
        'chart_json':     chart_json,
    })
