"""
OBR (Asosiy Zakaz) — bot bilan bir xil auto_zakaz.calculate_auto_zakaz() logikasi.

Yaxshilangan versiya:
- Kategoriya -> Podkategoriya -> Segment ierarxiyasi
- OBR% rang kodlash (yashil, sariq, to'q sariq, qizil)
- Qidiruv va filtrlar
- Sahifalash (pagination)
- Import sanasi filtri
"""
import json
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent))

from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.shortcuts import render
from django.http import JsonResponse
import openpyxl
import pandas as pd
import requests as _requests
import src.database.db_manager as db_manager
import auto_zakaz
import config as _cfg
from collections import defaultdict
from core.excel_export import style_header, style_total, set_col_widths, make_response


# ---------------------------------------------------------------------------
# Yordamchi funksiyalar
# ---------------------------------------------------------------------------

def _get_obr_df(request):
    """
    Session cache — bir sahifa so'rovida bir marta hisoblanadi.
    'refresh=1' parametri yuborilsa, keshni tozalab qayta hisoblaydi.
    """
    if request.GET.get('refresh'):
        request.session.pop('obr_df_json', None)

    if 'obr_df_json' not in request.session:
        df = auto_zakaz.calculate_auto_zakaz(db_manager.engine)
        if not df.empty:
            request.session['obr_df_json'] = df.to_json(
                orient='records', force_ascii=False
            )
        else:
            request.session['obr_df_json'] = ''

    raw = request.session.get('obr_df_json', '')
    if not raw:
        return pd.DataFrame()
    return pd.DataFrame(json.loads(raw))


def _obr_color_class(obr_value):
    """
    OBR foizi uchun CSS rang klassi qaytaradi.
    Bot logikasi bilan bir xil chegaralar:
      >= 99% -> yashil  (juda yaxshi aylanma)
      >= 85% -> sariq   (yaxshi)
      >= 70% -> to'q sariq (o'rtacha)
      < 70%  -> qizil   (past)
    """
    if obr_value >= 99:
        return 'obr-green'
    elif obr_value >= 85:
        return 'obr-yellow'
    elif obr_value >= 70:
        return 'obr-orange'
    else:
        return 'obr-red'


def _parse_obr_int(obr_str):
    """'145%' -> 145, xatolarni 0 deb oladi."""
    try:
        return int(str(obr_str).replace('%', '').strip())
    except (ValueError, TypeError):
        return 0


# ---------------------------------------------------------------------------
# Views
# ---------------------------------------------------------------------------

@login_required
def obr_root(request):
    """
    Bosh sahifa — barcha kategoriyalar ro'yxati + podkategoriya qidiruvi.
    """
    df = _get_obr_df(request)
    search_query = request.GET.get('q', '').strip()

    search_results = []
    if search_query and not df.empty:
        q_lower = search_query.lower()
        # Kategoriya nomi va podkategoriya nomida qidirish
        mask = (
            df['Подкатегория'].astype(str).str.lower().str.contains(q_lower, na=False) |
            df['Категория'].astype(str).str.lower().str.contains(q_lower, na=False) |
            df['Наименование'].astype(str).str.lower().str.contains(q_lower, na=False)
        )
        found = df[mask]
        # Kategoriya + Podkategoriya juftlarini topamiz (takrorlanmasin)
        pairs = found[['Категория', 'Подкатегория']].drop_duplicates()
        for _, row in pairs.iterrows():
            cat  = row['Категория']
            sub  = row['Подкатегория']
            sub_df = df[(df['Категория'] == cat) & (df['Подкатегория'] == sub)]
            total_zakaz = int(sub_df['Zakaz'].sum())
            zakaz_rows  = int((sub_df['Zakaz'] > 0).sum())
            total_rows  = len(sub_df)
            obr_vals    = sub_df['OBR %'].apply(_parse_obr_int)
            avg_obr     = int(obr_vals.mean()) if len(obr_vals) > 0 else 0
            search_results.append({
                'category':   cat,
                'subcategory': sub,
                'total_zakaz': total_zakaz,
                'zakaz_rows':  zakaz_rows,
                'total_rows':  total_rows,
                'avg_obr':    avg_obr,
                'obr_class':  _obr_color_class(avg_obr),
            })
        search_results.sort(key=lambda x: (x['category'], x['subcategory']))

    if df.empty:
        categories = []
    else:
        cat_stats = []
        for cat_name in sorted(df['Категория'].dropna().unique()):
            cat_df = df[df['Категория'] == cat_name]
            sub_count   = cat_df['Подкатегория'].nunique()
            total_zakaz = int(cat_df['Zakaz'].sum())
            total_rows  = len(cat_df)
            zakaz_rows  = int((cat_df['Zakaz'] > 0).sum())
            cat_stats.append({
                'name':        cat_name,
                'sub_count':   sub_count,
                'total_zakaz': total_zakaz,
                'total_rows':  total_rows,
                'zakaz_rows':  zakaz_rows,
            })
        categories = cat_stats

    # Autocomplete uchun barcha podkategoriyalar JSON
    ac_data = []
    if not df.empty:
        pairs = df[['Категория', 'Подкатегория']].drop_duplicates()
        for _, row in pairs.iterrows():
            ac_data.append({
                'label': f"{row['Подкатегория']}  ({row['Категория']})",
                'category':    row['Категория'],
                'subcategory': row['Подкатегория'],
            })
        ac_data.sort(key=lambda x: x['subcategory'])

    # -----------------------------------------------------------------
    # MAKRO TAHLIL — Kategoriya + Podkategoriya bo'yicha jins x segment
    # -----------------------------------------------------------------
    selected_macro_cat  = request.GET.get('macro_cat', '').strip()
    selected_macro_subs = [s.strip() for s in request.GET.getlist('macro_sub') if s.strip()]
    selected_macro_sub  = selected_macro_subs[0] if selected_macro_subs else ''  # backward compat
    macro_linked = request.GET.get('linked', '') == '1'  # default: bog'liq emas

    # Barcha kategoriyalar (filter uchun)
    all_cats = sorted(df['Категория'].dropna().unique().tolist()) if not df.empty else []

    if macro_linked and selected_macro_cat and not df.empty:
        # Bog'liq rejim: faqat tanlangan kategoriya podkategoriyalari
        macro_subcats = sorted(
            df[df['Категория'] == selected_macro_cat]['Подкатегория'].dropna().unique().tolist()
        )
    else:
        # Mustaqil rejim: barcha podkategoriyalar
        macro_subcats = sorted(df['Подкатегория'].dropna().unique().tolist()) if not df.empty else []

    macro_rows = []
    macro_totals = None
    macro_category = ''
    if selected_macro_subs and not df.empty:
        # Faqat podkategoriya bo'yicha filter (kategoriya ixtiyoriy)
        if selected_macro_cat:
            mdf = df[
                (df['Категория'] == selected_macro_cat) &
                (df['Подкатегория'].isin(selected_macro_subs))
            ].copy()
        else:
            mdf = df[df['Подкатегория'].isin(selected_macro_subs)].copy()

        # Category ni aniqlaymiz (link uchun)
        macro_category = selected_macro_cat if selected_macro_cat else (
            mdf['Категория'].iloc[0] if not mdf.empty else ''
        )

        if not mdf.empty:
            # Jins tartibi: Девочки -> Мальчики -> Универсал -> qolganlar
            pol_order = ['Девочки', 'Мальчики', 'Универсал']
            existing_pols = [p for p in pol_order if p in mdf['Пол'].values]
            for other in mdf['Пол'].dropna().unique():
                if other not in existing_pols:
                    existing_pols.append(other)

            grand_sotildi = grand_qoldiq = 0.0
            grand_zakaz = 0

            for pol in existing_pols:
                pol_df = mdf[mdf['Пол'] == pol].copy()
                pol_sotildi = float(pol_df['Продано'].sum())
                pol_qoldiq  = float(pol_df['Hozirgi_Qoldiq'].sum())
                pol_zakaz   = int(pol_df['Zakaz'].sum())
                pol_obr     = round(pol_sotildi / pol_qoldiq * 100) if pol_qoldiq > 0 else 0

                # Segmentlar — Jins + Segment bo'yicha GURUHLAB yig'amiz
                seg_grp = pol_df.groupby('Real_Sotuv_Segmenti', sort=True).agg(
                    sotildi=('Продано', 'sum'),
                    qoldiq=('Hozirgi_Qoldiq', 'sum'),
                    zakaz=('Zakaz', 'sum'),
                ).reset_index()

                for _, seg_row in seg_grp.iterrows():
                    seg     = seg_row['Real_Sotuv_Segmenti']
                    sotildi = float(seg_row['sotildi'])
                    qoldiq  = float(seg_row['qoldiq'])
                    zakaz   = int(seg_row['zakaz'])
                    obr_int = round(sotildi / qoldiq * 100) if qoldiq > 0 else 0
                    q_ulush = round(qoldiq / pol_qoldiq * 100) if pol_qoldiq > 0 else 0
                    s_ulush = round(sotildi / pol_sotildi * 100) if pol_sotildi > 0 else 0

                    macro_rows.append({
                        'type': 'row',
                        'pol': pol,
                        'segment': seg,
                        'obr': obr_int,
                        'sotildi': int(sotildi),
                        'qoldiq': int(qoldiq),
                        'q_ulush': q_ulush,
                        's_ulush': s_ulush,
                        'zakaz': zakaz,
                        'category': macro_category,
                    })

                # TOTAL qatori (jins bo'yicha)
                macro_rows.append({
                    'type': 'total',
                    'pol': f"{pol} TOTAL",
                    'segment': '',
                    'obr': pol_obr,
                    'sotildi': int(pol_sotildi),
                    'qoldiq': int(pol_qoldiq),
                    'q_ulush': 100,
                    's_ulush': 100,
                    'zakaz': pol_zakaz,
                    'category': macro_category,
                })

                grand_sotildi += pol_sotildi
                grand_qoldiq  += pol_qoldiq
                grand_zakaz   += pol_zakaz

            grand_obr = round(grand_sotildi / grand_qoldiq * 100) if grand_qoldiq > 0 else 0
            macro_totals = {
                'sotildi': int(grand_sotildi),
                'qoldiq': int(grand_qoldiq),
                'q_ulush': 100,
                's_ulush': 100,
                'zakaz': grand_zakaz,
                'obr': grand_obr,
            }

    return render(request, 'obr/root.html', {
        'categories':     categories,
        'search_query':   search_query,
        'search_results': search_results,
        'ac_data_json':   json.dumps(ac_data, ensure_ascii=False),
        'selected_macro_cat':  selected_macro_cat,
        'selected_macro_sub':  selected_macro_sub,   # eski nom saqlansin
        'selected_macro_subs': selected_macro_subs,
        'macro_sub_label':     ' + '.join(selected_macro_subs),
        'macro_linked':        macro_linked,
        'all_cats':            all_cats,
        'macro_subcats':       macro_subcats,
        'macro_rows':          macro_rows,
        'macro_totals':        macro_totals,
        'macro_category':      macro_category,
    })


@login_required
def obr_sub(request, category):
    """
    Tanlangan kategoriya ichidagi podkategoriyalar.
    Har bir podkategoriya uchun: segment soni, jami zakaz, tovar soni.
    """
    df = _get_obr_df(request)

    supplier_filter = request.GET.get('supplier', '').strip()

    if df.empty:
        subcategories = []
        cat_total_zakaz = cat_total_qoldiq = cat_total_sotuv = cat_total_obr = 0
    else:
        cat_df = df[df['Категория'] == category].copy()

        if supplier_filter:
            cat_df = cat_df[
                cat_df['Поставщик'].astype(str).str.strip().str.lower()
                == supplier_filter.lower()
            ]

        cat_total_zakaz = int(cat_df['Zakaz'].sum())
        cat_total_qoldiq = int(cat_df['Hozirgi_Qoldiq'].sum()) if 'Hozirgi_Qoldiq' in cat_df.columns else 0
        cat_total_sotuv  = int(cat_df['Продано'].sum()) if 'Продано' in cat_df.columns else 0
        cat_total_obr    = round(cat_total_sotuv / cat_total_qoldiq * 100) if cat_total_qoldiq > 0 else 0

        sub_stats = []
        for sub_name in sorted(cat_df['Подкатегория'].dropna().unique()):
            sub_df = cat_df[cat_df['Подкатегория'] == sub_name]
            seg_count = sub_df['Real_Sotuv_Segmenti'].nunique()
            total_zakaz = int(sub_df['Zakaz'].sum())
            total_rows = len(sub_df)
            zakaz_rows = int((sub_df['Zakaz'] > 0).sum())

            obr_vals = sub_df['OBR %'].apply(_parse_obr_int)
            avg_obr = int(obr_vals.mean()) if len(obr_vals) > 0 else 0

            sub_stats.append({
                'name': sub_name,
                'seg_count': seg_count,
                'total_zakaz': total_zakaz,
                'total_rows': total_rows,
                'zakaz_rows': zakaz_rows,
                'avg_obr': avg_obr,
                'obr_class': _obr_color_class(avg_obr),
            })
        subcategories = sub_stats

    if request.GET.get('export') == '1':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = category[:31]
        ws.append(['#', 'Podkategoriya', 'Segmentlar', 'Zakaz (dona)', 'Tovarlar', 'O\'rt OBR%'])
        style_header(ws, 1)
        for i, s in enumerate(subcategories, 1):
            ws.append([i, s['name'], s['seg_count'], s['total_zakaz'],
                       f"{s['zakaz_rows']}/{s['total_rows']}", f"{s['avg_obr']}%"])
        set_col_widths(ws, {'A': 5, 'B': 30, 'C': 12, 'D': 14, 'E': 12, 'F': 12})
        safe = category.replace('/', '_')
        return make_response(wb, f"obr_{safe}.xlsx")

    def _fmt(n):
        return f"{int(n):,}".replace(",", " ")

    return render(request, 'obr/sub.html', {
        'category':          category,
        'subcategories':     subcategories,
        'cat_total_zakaz':   cat_total_zakaz,
        'cat_total_qoldiq':  _fmt(cat_total_qoldiq),
        'cat_total_sotuv':   _fmt(cat_total_sotuv),
        'cat_total_obr':     cat_total_obr,
        'cat_obr_class':     _obr_color_class(cat_total_obr),
        'supplier_filter':   supplier_filter,
    })


@login_required
def obr_stat(request, category, subcategory):
    """
    Podkategoriya bo'yicha batafsil jadval.
    Filtrlash, qidiruv va sahifalash imkoniyati bilan.
    """
    df = _get_obr_df(request)

    if df.empty:
        orders = []
        total_packs = total_qoldiq = total_sotuv = total_obr = 0
        total_rows_count = 0
        total_obr_class = 'obr-red'
        page_obj = None
        search_query = ''
        show_all = False
        segment_filter = ''
        segments_list = []
        supplier_filter = ''
        suppliers_list = []
        pol_filter = ''
        pols_list = []
    else:
        sub_df = df[
            (df['Категория'] == category) &
            (df['Подкатегория'] == subcategory)
        ].copy()

        # OBR raqamini ajratib olish (sorting va filtrlash uchun)
        sub_df['obr_int'] = sub_df['OBR %'].apply(_parse_obr_int)

        # Kategoriya ichidagi barcha podkategoriyalar (dropdown uchun)
        sibling_subcategories = sorted(
            df[df['Категория'] == category]['Подкатегория'].dropna().unique().tolist()
        )

        # Barcha segmentlar, supplierlar va pollar (filtr uchun, filtering DAN OLDIN)
        segments_list = sorted(
            sub_df['Real_Sotuv_Segmenti'].dropna().unique().tolist()
        )
        suppliers_list = sorted(
            sub_df['Поставщик'].dropna().unique().tolist()
        )
        pols_list = sorted(
            sub_df['Пол'].dropna().unique().tolist()
        ) if 'Пол' in sub_df.columns else []

        # Autocomplete uchun barcha noyob nomlar (filtr qo'llanishidan OLDIN)
        names_json = json.dumps(
            sorted(sub_df['Наименование'].dropna().unique().tolist()),
            ensure_ascii=False
        )

        import re as _re

        def _norm(s):
            """Katta-kichik harf va bo'shliqni normallashtiradi."""
            return ' '.join(str(s).lower().split())

        # Ustun nomlari xavfsiz aniqlanadi (Материал2/Вид2 yo'q bo'lsa crash bo'lmasin)
        mat_col = 'Материал2' if 'Материал2' in sub_df.columns else (
            'Материал' if 'Материал' in sub_df.columns else None
        )
        vid_col = 'Вид2' if 'Вид2' in sub_df.columns else (
            'Вид' if 'Вид' in sub_df.columns else None
        )

        # --- Filtrlar (to'g'ri tartibda) ---

        # 1) Segment filtri — contains: "Bodi kr/r (14900)" → "1. Bodi kr/r (14900)" ham topadi.
        #    Agar filter hech narsaga mos kelmasa — filter QO'LLANMAYDI (hammasi qoladi),
        #    bo'sh natija ko'rsatilmaydi.
        segment_filter = request.GET.get('segment', '')
        if segment_filter:
            try:
                mask = sub_df['Real_Sotuv_Segmenti'].astype(str).str.contains(
                    segment_filter, regex=False, na=False
                )
                if mask.any():
                    sub_df = sub_df[mask]
                # else: mos kelmadi — filter o'tkazib yuboriladi (hammasi qoladi)
            except Exception:
                pass  # kutilmagan xato bo'lsa filter o'tkazib yuboriladi

        # 1.5) Pol (jins) filtri
        pol_filter = request.GET.get('pol', '')
        if pol_filter and 'Пол' in sub_df.columns:
            sub_df = sub_df[sub_df['Пол'] == pol_filter]

        # 2) Supplier filtri — case-insensitive, bo'shliq normalangan
        supplier_filter = request.GET.get('supplier', '')
        if supplier_filter:
            target = _norm(supplier_filter)
            sub_df = sub_df[sub_df['Поставщик'].apply(_norm) == target]

        # 3) Smart qidiruv — tab/ko'p bo'shliq normallanadi, tokenlar bo'yicha
        raw_q = request.GET.get('q', '')
        search_query = _re.sub(r'\s+', ' ', raw_q).strip()
        if search_query:
            tokens = search_query.lower().split()
            combined = sub_df['Наименование'].astype(str).str.lower()
            if mat_col:
                combined = combined + ' ' + sub_df[mat_col].astype(str).str.lower()
            combined = combined + ' ' + sub_df['Поставщик'].astype(str).str.lower()
            mask = combined.apply(lambda s: all(t in s for t in tokens))
            sub_df = sub_df[mask]

        # 4) show_all — ENG OXIRIDA (segment/supplier/search dan keyin)
        show_all = request.GET.get('show_all', '') == '1'
        if not show_all:
            sub_df = sub_df[sub_df['Zakaz'] > 0]

        # Saralash: OBR yuqori -> pastga
        sub_df = sub_df.sort_values('obr_int', ascending=False)

        # Nomi + Segment + Razmer + Jins bo'yicha group qilish (Supplier olib tashlanadi)
        import numpy as _np
        _group_cols = ['Наименование', 'Real_Sotuv_Segmenti', 'Размер сетка', 'Пол']
        # Faqat mavjud ustunlar
        _group_cols = [c for c in _group_cols if c in sub_df.columns]

        _agg = {}
        if 'Zakaz' in sub_df.columns:          _agg['Zakaz'] = 'sum'
        if 'Hozirgi_Qoldiq' in sub_df.columns: _agg['Hozirgi_Qoldiq'] = 'sum'
        if 'Продано' in sub_df.columns:        _agg['Продано'] = 'sum'
        if 'Ortacha_Qoldiq' in sub_df.columns: _agg['Ortacha_Qoldiq'] = 'mean'

        if _group_cols and _agg:
            sub_df = sub_df.groupby(_group_cols, as_index=False).agg(_agg).copy()
            # OBR qayta hisoblash
            if 'Продано' in sub_df.columns and 'Ortacha_Qoldiq' in sub_df.columns:
                sub_df['obr_int'] = (
                    sub_df['Продано'] / sub_df['Ortacha_Qoldiq'].replace(0, _np.nan)
                ).fillna(0).round(0).astype(int)
                sub_df['OBR %'] = sub_df['obr_int'].astype(str) + '%'
            elif 'obr_int' not in sub_df.columns:
                sub_df['obr_int'] = sub_df['OBR %'].apply(_parse_obr_int) if 'OBR %' in sub_df.columns else 0
            # Qayta saralash
            sub_df = sub_df.sort_values('obr_int', ascending=False)

        total_packs      = int(sub_df['Zakaz'].sum()) if 'Zakaz' in sub_df.columns else 0
        total_rows_count = len(sub_df)
        total_qoldiq     = int(sub_df['Hozirgi_Qoldiq'].sum()) if 'Hozirgi_Qoldiq' in sub_df.columns else 0
        total_sotuv      = int(sub_df['Продано'].sum()) if 'Продано' in sub_df.columns else 0
        total_obr        = round(total_sotuv / total_qoldiq * 100) if total_qoldiq > 0 else 0
        total_obr_class  = _obr_color_class(total_obr)

        # Jadval uchun dict ro'yxati
        orders = []
        for _, row in sub_df.iterrows():
            obr_val = int(row.get('obr_int', 0) or 0)
            orders.append({
                'naimenovanie': row.get('Наименование', ''),
                'segment': row.get('Real_Sotuv_Segmenti', ''),
                'razmer': row.get('Размер сетка', ''),
                'material': row.get(mat_col, '') if mat_col and mat_col in row.index else '',
                'vid': row.get(vid_col, '') if vid_col and vid_col in row.index else '',
                'pol': row.get('Пол', ''),
                'supplier': '',   # GROUP dan keyin supplier yo'q
                'zakaz': int(row.get('Zakaz', 0) or 0),
                'qoldiq': int(row.get('Hozirgi_Qoldiq', 0) or 0),
                'sotilgan': int(row.get('Продано', 0) or 0),
                'ortacha_qoldiq': int(row.get('Ortacha_Qoldiq', 0) or 0),
                'obr': str(obr_val) + '%',
                'obr_int': obr_val,
                'obr_class': _obr_color_class(obr_val),
            })

        # --- Sahifalash (Pagination) ---
        per_page = int(request.GET.get('per_page', 50))
        paginator = Paginator(orders, per_page)
        page_number = request.GET.get('page', 1)
        page_obj = paginator.get_page(page_number)

    if request.GET.get('export') == '1' and not df.empty:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = subcategory[:31]
        ws.freeze_panes = "B2"
        ws.append(['#', 'Nomi', 'Segment', 'Razmer', 'Pol', 'Supplier',
                   'Zakaz', 'Qoldiq', 'Sotilgan', 'O\'rt.Qoldiq', 'OBR%'])
        style_header(ws, 1)
        export_orders = orders  # already filtered (show_all applies)
        for i, o in enumerate(export_orders, 1):
            ws.append([i, o['naimenovanie'], o['segment'], o['razmer'],
                       o['pol'], o['supplier'], o['zakaz'], o['qoldiq'],
                       o['sotilgan'], o['ortacha_qoldiq'], o['obr']])
        ws.append(['', 'JAMI', '', '', '', '', total_packs, '', '', '', ''])
        style_total(ws, ws.max_row)
        set_col_widths(ws, {'A': 5, 'B': 35, 'C': 14, 'D': 12, 'E': 8,
                            'F': 20, 'G': 10, 'H': 10, 'I': 10, 'J': 12, 'K': 10})
        safe_cat = category.replace('/', '_')
        safe_sub = subcategory.replace('/', '_')
        return make_response(wb, f"obr_{safe_cat}_{safe_sub}.xlsx")

    return render(request, 'obr/stat.html', {
        'category': category,
        'subcategory': subcategory,
        'sibling_subcategories': sibling_subcategories if not df.empty else [],
        'names_json': names_json if not df.empty else '[]',
        'total_packs': total_packs,
        'total_rows_count': total_rows_count if df is not None and not df.empty else 0,
        'total_qoldiq':  total_qoldiq,
        'total_sotuv':   total_sotuv,
        'total_obr':     total_obr,
        'total_obr_class': total_obr_class,
        'page_obj': page_obj,
        'search_query': search_query if not df.empty else '',
        'show_all': show_all if not df.empty else False,
        'segment_filter': segment_filter if not df.empty else '',
        'segments_list': segments_list if not df.empty else [],
        'supplier_filter': supplier_filter if not df.empty else '',
        'suppliers_list': suppliers_list if not df.empty else [],
        'pol_filter': pol_filter if not df.empty else '',
        'pols_list': pols_list if not df.empty else [],
    })


@login_required
def obr_send_telegram(request, category, subcategory):
    """
    Tanlangan OBR qatorlarini Telegram arxiv kanaliga yuboradi.

    Frontend (stat.html) tanlangan checkboxlardagi data-* atributlarni
    JSON ko'rinishida POST qiladi. Bu yerda xabar belgilangan formatda
    yig'iladi va Telegram Bot API orqali jo'natiladi.
    """
    if request.method != 'POST':
        from django.http import HttpResponseNotAllowed
        return HttpResponseNotAllowed(['POST'])

    try:
        data = json.loads(request.body)
        items = data.get('items', [])
    except Exception:
        return JsonResponse({'ok': False, 'error': 'JSON xato'}, status=400)

    if not items:
        return JsonResponse({'ok': False, 'error': "Tanlangan element yo'q"}, status=400)

    pol_icons = {'Девочки': '👧', 'Мальчики': '👦', 'Универсал': '🌐'}

    # Guruhlanish: pol -> naimenovanie -> items (tartibni saqlab)
    groups = defaultdict(lambda: defaultdict(list))
    pol_order = []
    for item in items:
        pol = item.get('pol') or 'Универсал'
        naim = item.get('naimenovanie', '')
        if pol not in pol_order:
            pol_order.append(pol)
        if naim not in groups[pol]:
            # naimenovanie tartibini saqlash uchun bo'sh ro'yxat ochib qo'yamiz
            groups[pol][naim] = []
        groups[pol][naim].append(item)

    lines = [f"📦 {subcategory}\n"]
    for pol in pol_order:
        icon = pol_icons.get(pol, '🌐')
        lines.append(f"\n {icon} {pol}\n")
        n = 1
        for naim, naim_items in groups[pol].items():
            lines.append(f"\n 📌 {naim}")
            for item in naim_items:
                seg = item.get('segment', '')
                razm = item.get('razmer', '')
                mat = item.get('material', '')
                vid = item.get('vid', '')
                zak = item.get('zakaz', 0)
                sot = item.get('sotilgan', 0)
                qol = item.get('qoldiq', 0)
                obr = item.get('obr', '')
                lines.append(
                    f"\n{n}. {seg}|📐 {razm} |  {mat}  | 🎨 {vid} |\n"
                    f"  🛒 ZAKAZ: {zak} ta\n\n"
                    f"  | Sotuv: {sot} | Qoldiq: {qol} | OBR: {obr} |\n"
                    f"  {'─' * 22}"
                )
                n += 1

    message = "\n".join(lines)

    token = _cfg.TELEGRAM_BOT_TOKEN
    chat_id = _cfg.ARCHIVE_CHANNEL_ID

    if not token:
        return JsonResponse(
            {'ok': False, 'error': 'TELEGRAM_BOT_TOKEN sozlanmagan'}, status=500
        )

    # Telegram bitta xabar uchun 4096 belgi cheklovi bor — bo'laklarga ajratamiz.
    chunks = _split_message(message, 4000)

    try:
        for chunk in chunks:
            resp = _requests.post(
                f"https://api.telegram.org/bot{token}/sendMessage",
                json={"chat_id": chat_id, "text": chunk},
                timeout=15,
            )
            resp.raise_for_status()
        return JsonResponse({'ok': True, 'sent': len(items)})
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=500)


def _split_message(text, limit=4000):
    """
    Uzun xabarni Telegram cheklovidan oshmaydigan bo'laklarga ajratadi.
    Iloji boricha qator chegarasidan bo'ladi (formatni buzmaslik uchun).
    """
    if len(text) <= limit:
        return [text]

    chunks = []
    current = ''
    for line in text.split('\n'):
        candidate = line if not current else current + '\n' + line
        if len(candidate) > limit:
            if current:
                chunks.append(current)
            # Bitta qatorning o'zi limitdan uzun bo'lsa, qattiq kesamiz
            while len(line) > limit:
                chunks.append(line[:limit])
                line = line[limit:]
            current = line
        else:
            current = candidate
    if current:
        chunks.append(current)
    return chunks
